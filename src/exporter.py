"""
Export a mapped DataFrame to a styled .xlsx file with Hebrew RTL support.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HEADER_BG = "4472C4"
_HEADER_FG = "FFFFFF"
_HEADER_FONT_SIZE = 11
_DATA_FONT_SIZE = 10
_MIN_COL_WIDTH = 12
_MAX_COL_WIDTH = 48


def export_to_excel(df: pd.DataFrame, output_path: str) -> None:
    """Write *df* to *output_path* as a formatted Excel workbook."""
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    _style_header(ws)
    _style_data_rows(ws)
    _autosize_columns(ws)

    ws.sheet_view.rightToLeft = True
    wb.save(output_path)


def _style_header(ws) -> None:
    header_font = Font(bold=True, name="Arial", color=_HEADER_FG, size=_HEADER_FONT_SIZE)
    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    header_align = Alignment(
        horizontal="center", vertical="center",
        wrap_text=True, readingOrder=2,
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22


def _style_data_rows(ws) -> None:
    data_font = Font(name="Arial", size=_DATA_FONT_SIZE)
    data_align = Alignment(vertical="center", readingOrder=2)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align


def _autosize_columns(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value not in (None, "")),
            default=_MIN_COL_WIDTH,
        )
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
